import io

import pytest
from openpyxl import load_workbook

from app.errors import ForbiddenError, NotFoundError
from app.model.application import Application
from app.resources import report


def _identity(role="recruiter", comp_id="c1"):
    return {"id": "u1", "role": role, "comp_id": comp_id}


async def _seed(fakes, comp_id="c1", state="scored"):
    aid = await fakes["applications"].insert(
        Application(comp_id=comp_id, job_id="j1", candidate_user_id="cand", state=state)
    )
    fakes["reports"]._by_app[aid] = {
        "application_id": aid,
        "executive_summary": "Strong candidate",
        "highlights": ["clear communication"],
        "risks": ["limited scale experience"],
        "overall_score": 0.82,
        "recommendation": "advance",
    }
    return aid


async def test_get_report_returns_enriched(fakes):
    aid = await _seed(fakes)
    result = await report.get_report(
        _identity(), aid, applications=fakes["applications"], reports=fakes["reports"]
    )
    assert result["overall_score"] == 0.82
    assert result["recommendation"] == "advance"
    assert result["candidate_user_id"] == "cand"
    assert result["state"] == "scored"


async def test_get_report_carries_competency_evidence_and_integrity(fakes):
    aid = await fakes["applications"].insert(
        Application(comp_id="c1", job_id="j1", candidate_user_id="cand", state="scored")
    )
    fakes["reports"]._by_app[aid] = {
        "application_id": aid,
        "executive_summary": "Strong",
        "competency_scores": [
            {
                "competency": "python",
                "score": 0.9,
                "rationale": "solid",
                "evidence": [{"quote": "it yields control", "turn_index": 0}],
            }
        ],
        "integrity": {"score": 4.0, "flags": ["paste_large"], "auto_terminated": False},
        "overall_score": 0.8,
        "recommendation": "advance",
    }
    result = await report.get_report(
        _identity(), aid, applications=fakes["applications"], reports=fakes["reports"]
    )
    cs = result["competency_scores"][0]
    assert cs["competency"] == "python"
    assert cs["evidence"][0]["quote"] == "it yields control"
    assert result["integrity"]["flags"] == ["paste_large"]
    assert result["integrity"]["auto_terminated"] is False


async def test_get_report_defaults_missing_enrichment(fakes):
    # A pre-A4 report doc (no competency_scores/integrity) reads as empty/None — the DTO
    # is always well-formed so the FE never crashes on a legacy report.
    aid = await _seed(fakes)
    result = await report.get_report(
        _identity(), aid, applications=fakes["applications"], reports=fakes["reports"]
    )
    assert result["competency_scores"] == []
    assert result["integrity"] is None


async def test_get_report_rejects_other_company(fakes):
    aid = await _seed(fakes, comp_id="other")
    with pytest.raises(NotFoundError):
        await report.get_report(
            _identity(comp_id="c1"),
            aid,
            applications=fakes["applications"],
            reports=fakes["reports"],
        )


async def test_get_report_not_ready(fakes):
    aid = await fakes["applications"].insert(
        Application(
            comp_id="c1", job_id="j1", candidate_user_id="cand", state="interviewed"
        )
    )
    with pytest.raises(NotFoundError):
        await report.get_report(
            _identity(),
            aid,
            applications=fakes["applications"],
            reports=fakes["reports"],
        )


async def test_get_report_rejects_non_manager(fakes):
    aid = await _seed(fakes)
    with pytest.raises(ForbiddenError):
        await report.get_report(
            _identity(role="candidate"),
            aid,
            applications=fakes["applications"],
            reports=fakes["reports"],
        )


async def test_list_reports_for_job(fakes):
    aid = await _seed(fakes)
    results = await report.list_reports(
        _identity(), "j1", applications=fakes["applications"], reports=fakes["reports"]
    )
    assert len(results) == 1
    assert results[0]["application_id"] == aid
    assert results[0]["overall_score"] == 0.82


async def test_export_reports_xlsx(fakes):
    await _seed(fakes)
    content = await report.export_reports(
        _identity(), "j1", applications=fakes["applications"], reports=fakes["reports"]
    )
    ws = load_workbook(io.BytesIO(content)).active
    assert ws["A1"].value == "Candidate"  # header row
    assert ws["A2"].value == "cand"  # candidate_user_id
    assert ws["C2"].value == 0.82  # overall score
    assert ws["D2"].value == "advance"  # recommendation


async def test_list_reports_issues_single_batch_read(fakes):
    """AA-09 regression: list_reports must call list_by_applications once, not per-row.

    A serial per-application get_by_application loop would issue N round-trips;
    the batch path issues exactly one regardless of application count.
    """
    for _ in range(3):
        await _seed(fakes)
    await report.list_reports(
        _identity(), "j1", applications=fakes["applications"], reports=fakes["reports"]
    )
    assert len(fakes["reports"].list_by_applications_calls) == 1
    assert len(fakes["reports"].list_by_applications_calls[0]) == 3


async def test_list_reports_empty_job_returns_empty(fakes):
    """list_reports on a job with no applications returns an empty list."""
    result = await report.list_reports(
        _identity(),
        "empty-job",
        applications=fakes["applications"],
        reports=fakes["reports"],
    )
    assert result == []
